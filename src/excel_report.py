"""
excel_report.py  -  Render the FULL fact pack into a comprehensive, readable
Excel workbook so a human can review the analysis for completeness.

This is a visibility/review view of the fact pack (the agent's real input) - NOT
a second source of truth. If something is missing from the analysis, it should
be visibly missing here.

Tabs:
  Overview | By Region | By Zone | By Class | By ShopType |
  Tag-wise | Trend | Rankings | Data Quality | Guide
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = PatternFill("solid", fgColor="1F4E78")
SUB = PatternFill("solid", fgColor="D9E1F2")
TOTAL = PatternFill("solid", fgColor="FCE4D6")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
PCT2 = "0.00%"
NUM = "#,##0"
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _w(ws, widths):
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd


def _hdr(ws, row, names, fills=None):
    for i, n in enumerate(names, 1):
        c = ws.cell(row=row, column=i, value=n)
        c.fill = (fills[i - 1] if fills else BLUE)
        c.font = WHITE_BOLD
        c.alignment = CEN
        c.border = BORDER


def _cell(ws, r, c, v, fmt=None, fill=None, bold=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.border = BORDER
    if fmt and isinstance(v, (int, float)):
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if bold:
        cell.font = BOLD
    return cell


# ---------------------------------------------------------------------------
# Segment table (shared by Region / Zone / Class / ShopType)
# ---------------------------------------------------------------------------
SEG_COLS = [
    ("Group", "group", None),
    ("Outlets", "outlets", NUM),
    ("Booked outlets", "booked_outlets", NUM),
    ("Penetration", "penetration", PCT),
    ("Incremental value (est.)", "incremental_value", NUM),
    ("B: LPC growth", ("booked", "lpc_growth_ref1"), PCT),
    ("B: Value growth", ("booked", "value_growth_ref1"), PCT),
    ("B: Orders growth", ("booked", "orders_growth_ref1"), PCT),
    ("B: AOV growth", ("booked", "aov_growth_ref1"), PCT),
    ("B: % outlets up", ("booked", "lpc_up_rate"), PCT),
    ("B: Compliance", ("booked", "compliance"), PCT2),
    ("B: Qty conv", ("booked", "qty_conversion"), PCT2),
    ("B: Value conv", ("booked", "pr_value_conversion"), PCT2),
    ("B: Booked value", ("booked", "booked_value"), NUM),
    ("B: Missed value", ("booked", "missed_value"), NUM),
    ("NB: LPC growth", ("not_booked", "lpc_growth_ref1"), PCT),
    ("NB: Value growth", ("not_booked", "value_growth_ref1"), PCT),
    ("Lift: LPC", ("lift", "lpc"), PCT),
    ("Lift: Value", ("lift", "value"), PCT),
    ("Lift: Orders", ("lift", "orders"), PCT),
]


def _get(entry, path):
    if path is None:
        return None
    if isinstance(path, str):
        return entry.get(path)
    d = entry
    for k in path:
        d = (d or {}).get(k) if isinstance(d, dict) else None
    return d


def _segment_sheet(wb, title, rows, note):
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, color="555555")
    _hdr(ws, 2, [c[0] for c in SEG_COLS])
    r = 3
    for e in rows:
        for ci, (_, path, fmt) in enumerate(SEG_COLS, 1):
            _cell(ws, r, ci, _get(e, path), fmt)
        r += 1
    ws.freeze_panes = "B3"
    _w(ws, [22, 9, 11, 10, 16, 11, 11, 11, 11, 11, 11, 9, 10, 14, 14, 11, 11, 9, 9, 9])
    return ws


# ---------------------------------------------------------------------------
# Overview
# ---------------------------------------------------------------------------
def _overview(wb, fp):
    ws = wb.create_sheet("Overview")
    o = fp["overall"]; b = o.get("booked") or {}; nb = o.get("not_booked") or {}
    a = o.get("all") or {}; lift = o.get("lift") or {}
    cov = fp["coverage"]; dq = fp["data_quality"]; con = fp.get("concentration", {})
    tl = fp["timelines"]

    rows = [
        ("PR IMPACT - FULL ANALYSIS (Overview)", None, None),
        ("Source file", fp.get("source_file"), None),
        ("Direction", fp.get("direction"), None),
        ("", None, None),
        ("PERIODS", None, None),
    ]
    for k, v in tl.items():
        rows.append((f"  {v['label']}", f"{v['start']}..{v['end']}  ({v['days']}d)", None))
    rows += [
        ("", None, None),
        ("COVERAGE", None, None),
        ("Total outlets", cov.get("total_outlets"), NUM),
        ("Booked outlets", cov.get("booked_outlets"), NUM),
        ("Not-booked outlets", cov.get("not_booked_outlets"), NUM),
        ("Regions / Zones / Territories", f"{cov.get('regions')} / {cov.get('zones')} / {cov.get('territories')}", None),
        ("ShopTypes / Classes", f"{cov.get('shoptypes')} / {cov.get('classes')}", None),
        ("", None, None),
        ("DATA QUALITY (facts - your call, not a decision)", None, None),
        ("Usable base-vs-ref1", dq.get("usable_base_vs_ref1"), NUM),
        ("  as % of outlets", dq.get("pct_usable_base_vs_ref1"), PCT),
        ("Base LPC invalid", dq.get("base_lpc_invalid"), NUM),
        ("Ref1 LPC invalid", dq.get("ref1_lpc_invalid"), NUM),
        ("", None, None),
        ("OVERALL IMPACT - PR BOOKED (Ref1 vs Base)", None, None),
        ("LPC growth", b.get("lpc_growth_ref1"), PCT),
        ("Value growth", b.get("value_growth_ref1"), PCT),
        ("Orders growth", b.get("orders_growth_ref1"), PCT),
        ("AOV / drop-size growth", b.get("aov_growth_ref1"), PCT),
        ("% booked outlets improved (LPC)", b.get("lpc_up_rate"), PCT),
        ("Compliance (count)", b.get("compliance"), PCT2),
        ("Qty conversion", b.get("qty_conversion"), PCT2),
        ("Value conversion", b.get("pr_value_conversion"), PCT2),
        ("Booked value", b.get("booked_value"), NUM),
        ("Recommended value", b.get("rec_value"), NUM),
        ("Missed value (headroom)", b.get("missed_value"), NUM),
        ("Incremental value (est., PR-attributable)", fp.get("overall_incremental_value"), NUM),
        ("", None, None),
        ("OVERALL - NOT BOOKED (Ref1 vs Base)", None, None),
        ("LPC growth", nb.get("lpc_growth_ref1"), PCT),
        ("Value growth", nb.get("value_growth_ref1"), PCT),
        ("Orders growth", nb.get("orders_growth_ref1"), PCT),
        ("", None, None),
        ("LIFT (Booked − Not-booked, Ref1)", None, None),
        ("LPC lift", lift.get("lpc"), PCT),
        ("Value lift", lift.get("value"), PCT),
        ("Orders lift", lift.get("orders"), PCT),
        ("", None, None),
        ("CONCENTRATION (booked value across regions)", None, None),
        ("Top 5 regions share", con.get("top5_regions_share"), PCT),
        ("Top 10 regions share", con.get("top10_regions_share"), PCT),
    ]
    sections = {"PERIODS", "COVERAGE", "DATA QUALITY (facts - your call, not a decision)",
                "OVERALL IMPACT - PR BOOKED (Ref1 vs Base)", "OVERALL - NOT BOOKED (Ref1 vs Base)",
                "LIFT (Booked − Not-booked, Ref1)", "CONCENTRATION (booked value across regions)"}
    for i, (label, value, fmt) in enumerate(rows, 1):
        a_cell = ws.cell(row=i, column=1, value=label)
        v_cell = ws.cell(row=i, column=2, value=value)
        if i == 1:
            a_cell.font = Font(bold=True, size=14, color="1F4E78")
        elif label in sections:
            a_cell.font = BOLD; a_cell.fill = SUB; v_cell.fill = SUB
        else:
            a_cell.font = BOLD
        if fmt and isinstance(value, (int, float)):
            v_cell.number_format = fmt
    _w(ws, [42, 34])
    return ws


# ---------------------------------------------------------------------------
# Tag-wise
# ---------------------------------------------------------------------------
def _tags(wb, fp):
    ws = wb.create_sheet("Tag-wise")
    tags = fp.get("tag_wise_overall_booked") or {}
    _hdr(ws, 1, ["Tag", "Recommended value", "Booked value", "Conversion",
                 "Missed value", "Share of recommended", "Share of booked"])
    r = 2
    for t, v in tags.items():
        _cell(ws, r, 1, t.capitalize(), bold=True)
        _cell(ws, r, 2, v.get("recommended"), NUM)
        _cell(ws, r, 3, v.get("booked"), NUM)
        _cell(ws, r, 4, v.get("conversion"), PCT2)
        _cell(ws, r, 5, v.get("missed"), NUM)
        _cell(ws, r, 6, v.get("share_recommended"), PCT)
        _cell(ws, r, 7, v.get("share_booked"), PCT)
        r += 1
    _w(ws, [16, 18, 16, 12, 16, 20, 18])
    return ws


# ---------------------------------------------------------------------------
# Trend
# ---------------------------------------------------------------------------
def _tag_segment(wb, fp):
    ws = wb.create_sheet("Tag x Segment")
    ws.cell(row=1, column=1, value="Tag-wise booked conversion by segment (Upsell / Conventional / New / Focussed)").font = BOLD
    r = 2
    for title, key in [("By Zone", "tags_by_zone"), ("By Class", "tags_by_class")]:
        c = ws.cell(row=r, column=1, value=title); c.font = BOLD; c.fill = SUB
        r += 1
        _hdr(ws, r, ["Group", "Booked outlets", "Upsell", "Conventional", "New", "Focussed"])
        r += 1
        for e in fp.get(key, []):
            t = e.get("tags", {})
            _cell(ws, r, 1, e.get("group"))
            _cell(ws, r, 2, e.get("booked_outlets"), NUM)
            _cell(ws, r, 3, t.get("upsell"), PCT2)
            _cell(ws, r, 4, t.get("conventional"), PCT2)
            _cell(ws, r, 5, t.get("new"), PCT2)
            _cell(ws, r, 6, t.get("focussed"), PCT2)
            r += 1
        r += 1
    _w(ws, [24, 14, 12, 14, 10, 12])
    return ws


def _trend(wb, fp):
    ws = wb.create_sheet("Trend")
    _hdr(ws, 1, ["Reference", "Window (days)", "Short window?",
                 "All: LPC g", "All: Value g", "All: Orders g",
                 "Booked: LPC g", "Booked: Value g", "Booked: Orders g"])
    allt = {t["reference"]: t for t in fp.get("trend", [])}
    bkt = {t["reference"]: t for t in fp.get("trend_booked", [])}
    r = 2
    for ref in allt:
        a = allt[ref]; b = bkt.get(ref, {})
        _cell(ws, r, 1, ref, bold=True)
        _cell(ws, r, 2, a.get("days"), NUM)
        _cell(ws, r, 3, "YES" if a.get("short_window") else "no")
        _cell(ws, r, 4, a.get("lpc_growth"), PCT)
        _cell(ws, r, 5, a.get("value_growth"), PCT)
        _cell(ws, r, 6, a.get("orders_growth"), PCT)
        _cell(ws, r, 7, b.get("lpc_growth"), PCT)
        _cell(ws, r, 8, b.get("value_growth"), PCT)
        _cell(ws, r, 9, b.get("orders_growth"), PCT)
        r += 1
    _w(ws, [12, 14, 13, 11, 12, 12, 13, 14, 14])
    return ws


# ---------------------------------------------------------------------------
# Rankings
# ---------------------------------------------------------------------------
def _rankings(wb, fp):
    ws = wb.create_sheet("Rankings")
    rk = fp.get("rankings", {})
    blocks = [
        ("Region - best LIFT (LPC)", "region_best_lift", "lift_lpc", PCT),
        ("Region - worst LIFT (LPC)", "region_worst_lift", "lift_lpc", PCT),
        ("Region - top BOOKED VALUE (contribution)", "region_top_booked_value", "booked_value", NUM),
        ("Region - top MISSED VALUE (headroom)", "region_top_headroom", "missed_value", NUM),
        ("Region - best VALUE GROWTH", "region_best_value_growth", "value_growth_ref1", PCT),
        ("Region - worst VALUE GROWTH", "region_worst_value_growth", "value_growth_ref1", PCT),
        ("Territory - best LIFT (LPC)", "territory_best_lift", "lift_lpc", PCT),
        ("Territory - worst LIFT (LPC)", "territory_worst_lift", "lift_lpc", PCT),
    ]
    r = 1
    for title, key, field, fmt in blocks:
        c = ws.cell(row=r, column=1, value=title); c.font = BOLD; c.fill = SUB
        r += 1
        _hdr(ws, r, ["Group", "Booked outlets", field]); r += 1
        for item in rk.get(key, []):
            _cell(ws, r, 1, item.get("group"))
            _cell(ws, r, 2, item.get("booked_outlets"), NUM)
            _cell(ws, r, 3, item.get(field), fmt)
            r += 1
        r += 1
    _w(ws, [30, 14, 18])
    return ws


# ---------------------------------------------------------------------------
# Data Quality
# ---------------------------------------------------------------------------
def _dq(wb, fp):
    ws = wb.create_sheet("Data Quality")
    dq = fp["data_quality"]
    ws.cell(row=1, column=1, value="DATA QUALITY - facts only (whether acceptable is a human call)").font = Font(bold=True, color="1F4E78")
    items = [
        ("Total outlets", dq.get("total_outlets"), NUM),
        ("Usable for Base-vs-Ref1", dq.get("usable_base_vs_ref1"), NUM),
        ("  as %", dq.get("pct_usable_base_vs_ref1"), PCT),
        ("Base LPC invalid/zero", dq.get("base_lpc_invalid"), NUM),
        ("Ref1 LPC invalid/zero", dq.get("ref1_lpc_invalid"), NUM),
    ]
    r = 3
    for label, v, fmt in items:
        _cell(ws, r, 1, label, bold=True)
        _cell(ws, r, 2, v, fmt)
        r += 1
    _w(ws, [30, 18])
    return ws


# ---------------------------------------------------------------------------
# Guide
# ---------------------------------------------------------------------------
def _guide(wb):
    ws = wb.create_sheet("Guide")
    _hdr(ws, 1, ["Term", "Meaning"])
    guide = [
        ("Base / Reference", "Base = the newer month; Reference = the older month. Growth = (Base − Reference) / Reference."),
        ("LPC growth", "% change in Lines Per Call, Base vs Reference."),
        ("Value / Orders growth", "% change in gross value / order units, Base vs Reference."),
        ("Booked vs Not-booked", "Outlets that booked ≥1 recommendation vs those that didn't (the quasi-control)."),
        ("Lift", "Booked-group growth − Not-booked-group growth. The control-adjusted PR effect. Positive = booking helped."),
        ("Compliance", "PRs booked ÷ PRs given (count conversion of recommendations)."),
        ("Qty conversion", "Booked qty ÷ Recommended qty."),
        ("Value conversion", "Booked value ÷ Recommended value."),
        ("Missed value (headroom)", "Recommended value − Booked value - the unrealised opportunity still on the table."),
        ("Penetration", "Booked outlets ÷ total outlets in the group - how widely PR was adopted there."),
        ("Tag-wise", "Upsell / Conventional / New Product / Focussed - recommended vs booked value, conversion, and mix (share)."),
        ("Incremental value (est.)", "Estimated PR-attributable rupees = booked group's base value x value-lift. An estimate of the extra value PR drove versus the control."),
        ("AOV / drop-size", "Average order value = gross value ÷ orders. Its growth shows whether each order got bigger."),
        ("% outlets up", "Share of booked outlets whose LPC actually improved (Base > Ref1) - a distribution check beyond the average."),
        ("Tag x Segment", "Tag-wise booked conversion broken down by zone and by class - where each tag converts best/worst."),
        ("Concentration", "Share of total booked value sitting in the top few regions."),
        ("Trend", "Base compared against Reference 1, 2, 3 - recent movement. Short windows (<20 days) are flagged."),
        ("Data quality", "Counts of outlets usable/unusable for the comparison. Shown as facts - the judgement is yours."),
    ]
    r = 2
    for term, mean in guide:
        _cell(ws, r, 1, term, bold=True)
        c = _cell(ws, r, 2, mean); c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    _w(ws, [26, 95])
    return ws


# ---------------------------------------------------------------------------
def write_excel_report(fp, out_path, source_path=None):
    wb = Workbook()
    wb.remove(wb.active)
    _overview(wb, fp)
    _segment_sheet(wb, "By Region", fp.get("by_region", []), "Impact by Region (state). B = PR booked, NB = not booked.")
    _segment_sheet(wb, "By Zone", fp.get("by_zone", []), "Impact by Zone.")
    _segment_sheet(wb, "By Class", fp.get("by_class", []), "Impact by outlet Class (A-PLUS/A/B/C/D/Undefined).")
    _segment_sheet(wb, "By ShopType", fp.get("by_shoptype", []), "Impact by ShopType (channel).")
    _tags(wb, fp)
    _tag_segment(wb, fp)
    _trend(wb, fp)
    _rankings(wb, fp)
    _dq(wb, fp)
    _guide(wb)
    wb.save(out_path)
    return out_path
